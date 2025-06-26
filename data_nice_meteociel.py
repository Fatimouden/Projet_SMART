from meteociel.stations import station
from datetime import datetime

#extraction des données
date = datetime.strptime("2025/05/20", "%Y/%m/%d")
nice, data = station(date, "Nice")
print(data)

#enregistrer dans un fichier csv
fichier_nice_20_05_25_c=f"weather_{nice.lower()}_{date.strftime('%Y%m%d')}.csv"
data.to_csv(fichier_nice_20_05_25_c, index=False)
print(f"✅ Données sauvegardées dans : {fichier_nice_20_05_25_c}")

#enregistrer dans un fichier excel
fichier_nice_20_05_25_c_05_25_e = f"weather_{nice.lower()}_{date.strftime('%Y%m%d')}.xlsx"
data.to_excel(fichier_nice_20_05_25_e, index=False)
print(f"✅ Données enregistrées dans : {fichier_nice_20_05_25_e}")

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

#Charger le fichier Excel
wb = load_workbook(fichier_nice_20_05_25_e)
ws = wb.active

#Créer le graphique de température
chart = LineChart()
chart.title = f"Température à {nice} le {date.strftime('%d/%m/%Y')}"
chart.x_axis.title = "Heure"
chart.y_axis.title = "Température (°C)"

# Références : heure (X) et température (Y)
data_ref = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
time_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(time_ref)

#Ajouter le graphique à la feuille
ws.add_chart(chart, "G2")

#Sauvegarder
wb.save(fichier_nice_20_05_25_e)
print(f"✅ Graphique ajouté à : {fichier_nice_20_05_25_e}")



